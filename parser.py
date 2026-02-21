import logging
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, LongTable
from tqdm import tqdm
from yaspin import yaspin

from record_types import *


def to_para(contents, style):
    result = []
    for content in contents:
        result.append(Paragraph(content, style))
    return result


def replace_sub(text):
    if text == '':
        return 'N/A'
    else:
        return text


def log_time_msg(msg):
    t = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"{t} INFO  - {msg}"


def add_page_number(canvas, doc_instance):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.drawCentredString(A4[1] / 2, 0.5 * cm, f"Page {doc_instance.page}")
    canvas.restoreState()


logger = logging.getLogger()
logger.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
formatter = logging.Formatter(
    fmt='%(asctime)s %(levelname)-5s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        logger.info(f"Using .sss file: {file_path}")
    else:
        logger.critical("Usage: python parser.py <sss_file_path>")
        exit()

    f = open(file_path, 'rb')
    test_results: List[TestResult] = []
    machine_info = None

    # parsing
    while True:
        start_byte = f.read(1)
        assert start_byte == b'\x55'

        record_length = f.read(2)
        length_val = struct.unpack('<H', record_length)[0]
        logger.debug(f'Found record, length = {length_val}')

        checksum = f.read(2)
        checksum_val = struct.unpack('<H', checksum)[0]

        empty_byte = f.read(2)
        assert empty_byte == b'\x00\x00'

        # I don't know why but sometimes the length will short a little bit
        record_content = f.read(length_val)
        calculated_checksum = (sum(record_content)) & 0xffff
        if calculated_checksum == checksum_val or calculated_checksum == checksum_val + 1:
            logger.debug(f'Record checksum passed')
        else:
            length_val += 1
            record_content += f.read(1)
            calculated_checksum = (sum(record_content)) & 0xffff
            assert calculated_checksum == checksum_val or calculated_checksum == checksum_val + 1
            logger.debug(f'Record checksum passed')

        record_type = record_content[0]
        if record_type == 0xaa and record_content[1] == 0xff:
            logger.info(f"Parsed {len(test_results)} record, ready to write")
            break

        instance = record_type_class_defs[record_type](record_content[1:])
        if type(instance) is MachineInfo:
            machine_info = instance
        else:
            test_results.append(instance)
            test_results.append(instance)
        logger.debug(f'Record content = {instance}')

    result_file_path = f"{os.path.splitext(os.path.basename(file_path))[0]}_parsed_{datetime.now().strftime('%y_%m_%d_%H_%M_%S')}"
    logger.info(f'Writing into Excel+PDF file...')

    # %% Excel preparation
    wb = Workbook()
    ws = wb.active
    fill = PatternFill(start_color="F56C6C", end_color="F56C6C", fill_type="solid")

    # Instrument info
    ws.append(
        ["Test Instrument Model", "", "", machine_info.machine_model, "", "", "", "Test Instrument Serial Number", "",
         "", machine_info.machine_serial_number, "", "", ""])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
    ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=14)

    # headers
    headers = [
        "Asset ID", "Site Name", "Location Name", "Test Time",
        "Test Operator", "Overall Result", "Program", "Comments",
        "Next Full Test Date", "Next Formal Visual Test Date", "Test Result"
    ]
    ws.append(headers)

    # sub headers
    sub_headers = ["", "", "", "", "", "", "", "", "", "", "Test Type", "Result", "Unit", "Status"]
    ws.append(sub_headers)

    for col in range(1, 11):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column=14)

    # %% PDF preparation
    doc = SimpleDocTemplate(f"{result_file_path}.pdf",
                            title='Portable Appliance Test (PAT) Report',
                            pagesize=landscape(A4),
                            rightMargin=1 * cm, leftMargin=1 * cm,
                            topMargin=1 * cm, bottomMargin=1 * cm)
    elements = []
    styles = getSampleStyleSheet()

    style_section_header = ParagraphStyle(name='SectionHeader', parent=styles['Normal'], fontSize=9,
                                          textColor=colors.white, fontName='Helvetica-Bold')
    style_normal = ParagraphStyle(name='NormalText', parent=styles['Normal'], fontSize=8, fontName='Helvetica')
    style_normal_centered = ParagraphStyle(name='NormalText', parent=styles['Normal'], fontSize=8, fontName='Helvetica',
                                           alignment=TA_CENTER)
    highlight_size = ParagraphStyle(name='normal_size_bold', parent=styles['Normal'], fontSize=10, fontName='Helvetica')

    dark_blue = colors.Color(0.1, 0.2, 0.4)
    light_grey = colors.Color(0.95, 0.95, 0.95)

    header_row_style = [
        ('BACKGROUND', (0, 0), (-1, 0), dark_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ]

    elements.append(Paragraph("Portable Appliance Test (PAT) Report", ParagraphStyle(
        name='HeaderTitle',
        parent=styles['Normal'],
        fontSize=16,
        fontName='Helvetica-Bold',
        leftIndent=-0.25 * cm
    )))
    elements.append(Spacer(1, 1 * cm))

    # tec info
    tec_info_table_content = [
        [Paragraph("TESTING CARRIED OUT BY", style_section_header)],
        [Paragraph("<b>TEC PA & Lighting</b>", highlight_size)],
        [Paragraph("""<b>Email:</b> info@nottinghamtec.co.uk<br/>
            <b>Website:</b> www.nottinghamtec.co.uk<br/>
            <b>Tel:</b> 0115 84 68720<br/>
            <b>Address:</b><br/>Portland Building<br/>University Park<br/>Nottingham<br/>NG7 2RD<br/>""", style_normal)]
    ]
    tec_info_table = Table(tec_info_table_content, colWidths=[doc.width / 2 - 1 * cm], )
    tec_info_table.setStyle(TableStyle(header_row_style + [
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
    ]))

    tec_logo = Image("imgs/combined-logo.png")
    w, h = tec_info_table.wrap(doc.width / 2 - 1 * cm, doc.height)
    target_height = h * 0.8
    aspect_ratio = tec_logo.imageWidth / tec_logo.imageHeight
    tec_logo.drawHeight = target_height
    tec_logo.drawWidth = target_height * aspect_ratio

    tec_info = Table(
        [[tec_info_table, tec_logo]],
        colWidths=[doc.width / 2, doc.width / 2],
        hAlign='CENTER'
    )
    tec_info.setStyle(TableStyle([
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('VALIGN', (0, 0), (0, 0), 'TOP'),
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))

    elements.append(tec_info)
    elements.append(Spacer(1, 0.5 * cm))

    # tester info
    tester_info_content = [
        [Paragraph("PAT TESTER INFO", style_section_header), ""],
        ["Serial Number", "Make and Model"],
        [machine_info.machine_serial_number, machine_info.machine_model],
    ]
    tester_info_table = Table(tester_info_content, colWidths=[doc.width / 2, doc.width / 2])
    tester_info_table.setStyle(TableStyle(
        header_row_style +
        [
            ('SPAN', (0, 0), (1, 0))
        ] +
        [
            ('GRID', (0, 1), (-1, -1), 0.5, colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, 1), light_grey),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ] + [
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        ]
    ))
    elements.append(tester_info_table)
    elements.append(Spacer(1, 0.5 * cm))

    # test result table header
    result_header_content = [
        [Paragraph("APPLIANCE DETAILS AND TEST RESULTS", style_section_header)],
        [Paragraph("<b>Key</b><br/>PASS / FAIL / INFO / N/A = Not Applicable", style_normal)],
    ]
    result_header_table = Table(result_header_content, colWidths=[doc.width])
    result_header_table.setStyle(TableStyle(
        header_row_style +
        [
            ('BACKGROUND', (0, 1), (-1, -1), light_grey),
            ('GRID', (0, 1), (-1, -1), 0.5, colors.lightgrey),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.grey),
            ('LINEBEFORE', (0, 0), (0, -1), 1, colors.grey),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.grey),
        ]
    ))
    elements.append(result_header_table)

    # data
    result_header_to_be_repeated = [
        ['Appliance ID', 'Appliance Description', 'Test Date', 'Operator', 'Program', 'Test Items', '', "", "",
         "Overall Status", 'Comments'],
        ['', '', '', '', '', 'Test Type', 'Result', "Unit", "Status", "", ''],
    ]
    result_content = []
    result_style = []

    # show records in different site/location seperately
    locations = []
    record_grouped_by_location = {}
    for record in test_results:
        location = f"{record.site_name} - {record.location_name}"
        if location not in locations:
            locations.append(location)
            record_grouped_by_location[location] = []
        record_grouped_by_location[location].append(record)

    current_row_excel = 4
    current_row_pdf = 2
    with yaspin(text="Formatting Result", color="black") as spinner:
        with tqdm(total=sum(len(l) for _, l in record_grouped_by_location.items()), desc="Progress", position=1,
                  leave=False,
                  bar_format="{l_bar} {bar}| {n}/{total}") as pbar:
            for location, records in record_grouped_by_location.items():
                # add a placeholder for site title because also want to add failed test number
                header_idx = len(result_content)
                result_content.append('')
                result_style.extend([
                    ('BACKGROUND', (0, current_row_pdf), (-1, current_row_pdf), dark_blue),
                    ('TEXTCOLOR', (0, current_row_pdf), (-1, current_row_pdf), colors.white),
                    ('SPAN', (0, current_row_pdf), (-1, current_row_pdf)),
                ])
                current_row_pdf += 1
                failed_counter = 0

                for record in records:
                    formatted_results = []
                    used_row = 0
                    merge_row = []
                    failed_row = []
                    original_result_style_len = len(result_style)

                    for visual_test_result in record.visual_test_results:
                        formatted_results.append([
                            visual_test_result.name,
                            visual_test_result.result if visual_test_result.unit else "",
                            visual_test_result.unit,
                            visual_test_result.flags[0]
                        ])
                        if not visual_test_result.unit:
                            merge_row.append(current_row_excel + used_row)
                        if visual_test_result.flags[0] == 'FAIL':
                            failed_row.append(current_row_excel + used_row)
                            result_style.append(
                                ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                 colors.Color(0.99, 0.88, 0.88))
                            )

                        used_row += 1
                    for physical_test_result in record.physical_test_results:
                        if isinstance(physical_test_result, EarthResistanceTestResult):
                            formatted_results.append([
                                "Earth Continuity",
                                physical_test_result.get_value(),
                                physical_test_result.resistance.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            used_row += 1
                        elif isinstance(physical_test_result, IECLeadContinuityTestResult):
                            formatted_results.append([
                                "IEC Lead Continuity",
                                physical_test_result.get_value(),
                                physical_test_result.resistance.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            used_row += 1
                        elif isinstance(physical_test_result, PointToPointTestResult):
                            formatted_results.append([
                                "Point To Point Resistance",
                                physical_test_result.get_value(),
                                physical_test_result.resistance.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            used_row += 1
                        elif isinstance(physical_test_result, InsulationTestResult):
                            formatted_results.append([
                                "Insulation",
                                physical_test_result.get_value(),
                                physical_test_result.resistance.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            formatted_results.append([
                                "Insulation Voltage",
                                physical_test_result.voltage.value,
                                physical_test_result.voltage.unit,
                                "INFO"
                            ])
                            used_row += 2
                        elif isinstance(physical_test_result, SubstituteLeakageTestResult):
                            formatted_results.append([
                                "Substitute Leakage Current",
                                physical_test_result.get_value(),
                                physical_test_result.current.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            used_row += 1
                        elif isinstance(physical_test_result, PolarityTestResult):
                            merge_row.append(current_row_excel + used_row)
                            formatted_results.append([
                                "IEC Lead Polarity",
                                physical_test_result.get_value(),
                                "",
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            used_row += 1
                        elif isinstance(physical_test_result, MainVoltageTestResult):
                            formatted_results.append([
                                "Main Voltage",
                                physical_test_result.get_value(),
                                physical_test_result.voltage.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row),
                                     colors.Color(0.99, 0.88, 0.88))
                                )

                            used_row += 1
                        elif isinstance(physical_test_result, TouchOrLeakageCurrentTestResult):
                            formatted_results.append([
                                "Touch Or Leakage Test Load Current",
                                physical_test_result.load_current.value,
                                physical_test_result.load_current.unit,
                                physical_test_result.get_status()
                            ])
                            formatted_results.append([
                                "Touch Or Leakage Test Leakage Current",
                                physical_test_result.leakage_current.value,
                                physical_test_result.leakage_current.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(current_row_excel + used_row)
                                failed_row.append(current_row_excel + used_row + 1)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row), (8, current_row_pdf + used_row + 1),
                                     colors.Color(0.99, 0.88, 0.88))
                                )
                            used_row += 2
                        elif isinstance(physical_test_result, RCDTestResult):
                            formatted_results.append([
                                "RCD Test Current",
                                physical_test_result.test_current.value,
                                physical_test_result.test_current.unit,
                                "INFORMATION"
                            ])
                            formatted_results.append([
                                "RCD Test Circle Angle",
                                physical_test_result.circle_angle.value,
                                physical_test_result.circle_angle.unit,
                                "INFORMATION"
                            ])
                            formatted_results.append([
                                "RCD Test Trip time",
                                physical_test_result.get_value(),
                                physical_test_result.trip_time.unit,
                                physical_test_result.get_status()
                            ])
                            if physical_test_result.get_status() == 'FAIL':
                                failed_row.append(used_row + 2)
                                result_style.append(
                                    ('BACKGROUND', (5, current_row_pdf + used_row + 2),
                                     (8, current_row_pdf + used_row + 2),
                                     colors.Color(0.99, 0.88, 0.88))
                                )
                            used_row += 3
                        elif isinstance(physical_test_result, StringComment):
                            merge_row.append(current_row_excel + used_row)
                            formatted_results.append([
                                physical_test_result.string_value,
                                "",
                                "",
                                physical_test_result.get_status()
                            ])
                            used_row += 1

                    if not len(result_style) == original_result_style_len:
                        result_style.extend(
                            [
                                ('BACKGROUND', (0, current_row_pdf), (0, current_row_pdf + used_row - 1),
                                 colors.Color(0.99, 0.88, 0.88)),
                                ('BACKGROUND', (9, current_row_pdf), (9, current_row_pdf + used_row - 1),
                                 colors.Color(0.99, 0.88, 0.88)),
                            ]
                        )
                        failed_counter += 1

                    for i, tr in enumerate(formatted_results):
                        tr[0] = replace_sub(str(tr[0]))
                        tr[1] = replace_sub(str(tr[1]))
                        tr[2] = replace_sub(str(tr[2]).lower())
                        tr[3] = replace_sub(str(tr[3]))

                        # excel
                        row = []
                        if i == 0:
                            row.extend([
                                record.asset_id,
                                record.site_name,
                                record.location_name,
                                record.test_time,
                                record.test_operator,
                                record.get_status(),
                                record.program,
                                record.comments,
                                record.next_full_test_date,
                                record.next_formal_visual_test_date
                            ])
                        else:
                            row.extend([""] * 10)

                        # pdf
                        row.extend(tr)
                        ws.append(row)

                        row = []
                        if i == 0:
                            row.extend([
                                record.asset_id,
                                '',
                                record.test_time.strftime("%d/%m/%Y"),
                                record.test_operator,
                                record.program,
                            ])
                            row.extend(tr)
                            row.extend([
                                record.get_status(),
                                record.comments
                            ])
                        else:
                            row.extend([""] * 5)
                            row.extend(tr)
                            row.extend([""] * 2)

                        result_content.append(to_para(row, style_normal_centered))

                    # merge information cols
                    if used_row > 1:
                        for col in range(1, 11):
                            ws.merge_cells(
                                start_row=current_row_excel,
                                start_column=col,
                                end_row=current_row_excel + used_row - 1,
                                end_column=col
                            )
                    for i in merge_row:
                        ws.merge_cells(
                            start_row=i,
                            start_column=12,
                            end_row=i,
                            end_column=13
                        )

                    # highlight fail cells
                    if len(failed_row) > 0:
                        ws.cell(row=current_row_excel, column=1).fill = fill
                        for row in failed_row:
                            for col in range(11, 15):
                                ws.cell(row=row, column=col).fill = fill

                    # pdf, add divider lines
                    for i in range(current_row_pdf, current_row_pdf + used_row):
                        result_style.extend([
                            ('LINEBELOW', (5, i), (8, i), 0.5, colors.lightgrey),
                        ])
                    result_style.extend([
                        ('LINEBELOW', (0, current_row_pdf + used_row - 1), (-1, current_row_pdf + used_row - 1), 1,
                         colors.grey),
                    ])

                    current_row_excel += used_row
                    current_row_pdf += used_row
                    pbar.update(1)

                # update placeholder created before
                result_content[header_idx] = to_para(
                    [f"{location} ({len(records)} Records in total, {failed_counter} FAILED)"] + [''] * 10,
                    style_section_header)

        spinner.ok(log_time_msg("✔"))

    with yaspin(text="Generating Excel File...", color="black") as spinner:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(f"{result_file_path}.xlsx")
        spinner.ok(log_time_msg("✔"))

    result_table = LongTable(result_header_to_be_repeated + result_content, colWidths=[
        3 * cm,
        6 * cm,
        2 * cm,
        2 * cm,
        2.5 * cm,
        3 * cm, 1.75 * cm, 1.25 * cm, 1.5 * cm,
        2.5 * cm,
        2.2 * cm
    ], repeatRows=2)
    result_table.setStyle(TableStyle(
        [
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, 1), 0.5, colors.lightgrey),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.grey),
            ('LINEBEFORE', (0, 0), (0, -1), 1, colors.grey),
            ('LINEAFTER', (10, 0), (10, -1), 1, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW', (0, 1), (-1, 1), 1, colors.grey),

            ('LINEAFTER', (0, 2), (0, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (1, 2), (1, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (2, 2), (2, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (3, 2), (3, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (4, 2), (4, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (5, 2), (5, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (6, 2), (6, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (7, 2), (7, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (8, 2), (8, -1), 0.5, colors.lightgrey),
            ('LINEAFTER', (9, 2), (9, -1), 0.5, colors.lightgrey),

            ('SPAN', (5, 0), (8, 0)),
            ('SPAN', (0, 0), (0, 1)),
            ('SPAN', (1, 0), (1, 1)),
            ('SPAN', (2, 0), (2, 1)),
            ('SPAN', (3, 0), (3, 1)),
            ('SPAN', (4, 0), (4, 1)),
            ('SPAN', (9, 0), (9, 1)),
            ('SPAN', (10, 0), (10, 1)),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 1), light_grey),
        ] + result_style
    ))
    elements.append(result_table)

    with yaspin(text="Generating PDF File...", color="black") as spinner:
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        spinner.ok(log_time_msg("✔"))

    logger.info(
        f"All test results have been written to {result_file_path}.xlsx/pdf, total records = {len(test_results)}")
